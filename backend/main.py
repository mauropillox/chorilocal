from fastapi import FastAPI, HTTPException, Depends, Form, Request, Query, APIRouter
from fastapi.middleware.cors import CORSMiddleware
# from starlette.middleware.gzip import GZIPMiddleware
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.responses import Response
from pydantic import BaseModel, field_validator, Field
from typing import List, Optional
from datetime import datetime, timedelta, timezone
from jose import JWTError, jwt
import os
from passlib.context import CryptContext
import db
import pdf_utils
import shutil
from uuid import uuid4
from fastapi import File, UploadFile
from fastapi.staticfiles import StaticFiles
from slowapi import Limiter
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from fastapi.responses import JSONResponse
import threading
import time
import urllib.request
import re

# API Version
API_VERSION = "1.0.0"

# Configuration from environment variables with sensible defaults
KEEP_ALIVE_INTERVAL_SECONDS = int(os.getenv("KEEP_ALIVE_INTERVAL", "600"))  # 10 minutes
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.getenv("ACCESS_TOKEN_EXPIRE_MINUTES", "60"))
MAX_UPLOAD_SIZE_MB = int(os.getenv("MAX_UPLOAD_SIZE_MB", "5"))
RATE_LIMIT_DEFAULT = os.getenv("RATE_LIMIT_DEFAULT", "100/minute")
RATE_LIMIT_LOGIN = os.getenv("RATE_LIMIT_LOGIN", "5/minute")

# Self-ping to keep Render free tier awake
def keep_alive():
    """Ping self every N minutes to prevent Render cold starts"""
    SELF_URL = os.getenv("RENDER_EXTERNAL_URL", os.getenv("SELF_URL", ""))
    if not SELF_URL:
        return  # Only run in production with URL configured
    
    health_url = f"{SELF_URL}/health"  # Correct: /health not /api/health
    while True:
        time.sleep(KEEP_ALIVE_INTERVAL_SECONDS)
        try:
            urllib.request.urlopen(health_url, timeout=10)
        except Exception:
            pass  # Ignore errors, just keep trying

# Start keep-alive thread in production
if os.getenv("ENVIRONMENT") == "production" or os.getenv("RENDER"):
    keep_alive_thread = threading.Thread(target=keep_alive, daemon=True)
    keep_alive_thread.start()

def _is_valid_image_url(url: Optional[str]) -> bool:
    if not url:
        return True
    try:
        from urllib.parse import urlparse
        p = urlparse(url)
        # allow full http(s) URLs or relative media paths (/media/...)
        if p.scheme in ("http", "https") and bool(p.netloc):
            return True
        if url.startswith('/media/'):
            return True
        return False
    except Exception:
        return False

# === Configuración ===
import logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Security: SECRET_KEY is required in production
SECRET_KEY = os.getenv("SECRET_KEY")
if not SECRET_KEY:
    if os.getenv("ENVIRONMENT", "development") == "production":
        raise ValueError("SECRET_KEY environment variable is required in production!")
    # Generate random key for development (not persistent across restarts)
    import secrets
    SECRET_KEY = secrets.token_hex(32)
    logger.warning("⚠️  Using random development SECRET_KEY. Set SECRET_KEY env var for production!")


# === Helper para contexto de auditoría ===
def get_request_context(request: Request) -> dict:
    """Extrae IP y User-Agent del request para auditoría"""
    forwarded = request.headers.get("X-Forwarded-For")
    ip = forwarded.split(",")[0].strip() if forwarded else request.client.host if request.client else "unknown"
    user_agent = request.headers.get("User-Agent", "")[:200]
    return {"ip": ip, "user_agent": user_agent}

ALGORITHM = "HS256"
# ACCESS_TOKEN_EXPIRE_MINUTES is now defined at top of file from env

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="login")

db.verificar_tablas_y_columnas()
db.ensure_indexes()  # Add performance indexes
db.seed_ofertas_demo()  # Seed demo offers
db.seed_categorias_default()  # Seed default categories

# Cleanup expired revoked tokens on startup
try:
    cleaned = db.cleanup_expired_tokens()
    if cleaned > 0:
        logger.info(f"Cleaned up {cleaned} expired revoked tokens on startup")
except Exception as e:
    logger.warning(f"Failed to cleanup expired tokens: {e}")

app = FastAPI(
    title="Chorizaurio API",
    description="Gestión de pedidos y clientes de carnicería",
    version="1.0.0",
    docs_url="/docs",
    redoc_url="/redoc",
    openapi_url="/openapi.json"
)

# Initialize rate limiter
limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter

# Rate limit exceeded handler - return 429 instead of 500
@app.exception_handler(RateLimitExceeded)
async def rate_limit_handler(request: Request, exc: RateLimitExceeded):
    return JSONResponse(
        status_code=429,
        content={"detail": f"Rate limit exceeded: {exc.detail}"},
        headers={"Retry-After": str(getattr(exc, 'retry_after', 60))}
    )

# Add Gzip compression (responses -70% size)
try:
    from starlette.middleware.gzip import GZIPMiddleware
    app.add_middleware(GZIPMiddleware, minimum_size=1000)
except ImportError:
    try:
        from starlette.middleware import GZipMiddleware
        app.add_middleware(GZipMiddleware, minimum_size=1000)
    except ImportError:
        logger.warning("GZIPMiddleware not available, skipping compression")

# Security: CORS with specific origins
CORS_ORIGINS = os.getenv("CORS_ORIGINS", "http://localhost,http://localhost:80,http://localhost:8000").split(",")
# Allow all in development, specific origins in production
if os.getenv("ENVIRONMENT") == "production" and "*" in CORS_ORIGINS:
    logger.warning("⚠️  CORS is set to allow all origins in production. Configure CORS_ORIGINS env var!")

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type", "Accept"]
)

# Security: HSTS header for HTTPS enforcement in production
@app.middleware("http")
async def add_security_headers(request: Request, call_next):
    response = await call_next(request)
    if os.getenv("ENVIRONMENT") == "production" or os.getenv("RENDER"):
        # Enforce HTTPS for 1 year, include subdomains
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
    return response

# Ensure uploads dir exists on startup
UPLOAD_DIR = os.getenv('UPLOAD_DIR', '/data/uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)

# Security: File upload constraints
MAX_UPLOAD_SIZE = 5 * 1024 * 1024  # 5MB max
ALLOWED_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.pdf', '.txt'}
ALLOWED_MIME_TYPES = {
    'image/jpeg', 'image/png', 'image/gif', 'image/webp',
    'application/pdf', 'text/plain'
}

# Serve uploaded media at /media (maps to /data)
MEDIA_DIR = os.getenv('MEDIA_DIR', os.getenv('UPLOAD_DIR', '/data'))
app.mount('/media', StaticFiles(directory=MEDIA_DIR), name='media')

# API Versioning: Create v1 router for future-proofing
# All new routes should use api_v1 router
# Legacy routes remain on root for backward compatibility
api_v1 = APIRouter(prefix="/v1", tags=["v1"])

# === Helpers ===
def hash_password(password: str):
    return pwd_context.hash(password)

def verify_password(password: str, hash: str):
    return pwd_context.verify(password, hash)

def create_access_token(data: dict, expires_delta: timedelta = None):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + (expires_delta or timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES))
    # Add unique token ID (jti) for revocation support
    jti = str(uuid4())
    to_encode.update({"exp": expire, "jti": jti, "exp_iso": expire.isoformat()})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

def get_current_user(token: str = Depends(oauth2_scheme)):
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username = payload.get("sub")
        rol = payload.get("rol")
        jti = payload.get("jti")
        
        # Atomic check: user active AND token not revoked (eliminates TOCTOU race condition)
        if jti:
            user_db = db.get_active_user_if_token_valid(username, jti)
            if not user_db:
                # Could be revoked token OR inactive user - check which for logging
                if db.is_token_revoked(jti):
                    logger.warning(f"Revoked token used by: {username}")
                    raise HTTPException(status_code=401, detail="Sesión cerrada. Inicia sesión nuevamente.")
                else:
                    logger.warning(f"Token for inactive/deleted user: {username}")
                    raise HTTPException(status_code=401, detail="Tu cuenta ha sido desactivada. Contacta al administrador.")
        else:
            # Legacy token without jti - fallback to old behavior
            user_db = db.get_usuario(username)
            if not user_db or not user_db.get("activo"):
                logger.warning(f"Token for inactive/deleted user: {username}")
                raise HTTPException(status_code=401, detail="Tu cuenta ha sido desactivada. Contacta al administrador.")
        
        user = {"username": username, "rol": rol}
        logger.info(f"Auth: user={user['username']}, rol={user['rol']}")
        return user
    except JWTError as e:
        logger.warning(f"Invalid token attempt")
        raise HTTPException(status_code=401, detail="Token inválido o expirado. Inicia sesión nuevamente.")

def get_admin_user(user=Depends(get_current_user)):
    if user.get("rol") != "admin":
        logger.warning(f"Unauthorized access: user={user.get('username')}, rol={user.get('rol')}")
        raise HTTPException(status_code=403, detail="Acceso denegado. Solo administradores pueden realizar esta acción.")
    return user

def require_role(allowed_roles: list):
    def check_role(user=Depends(get_current_user)):
        if user.get("rol") not in allowed_roles:
            logger.warning(f"Unauthorized access: user={user.get('username')}, rol={user.get('rol')}")
            raise HTTPException(status_code=403, detail="Permiso denegado")
        return user
    return check_role

# === Modelos ===
class Cliente(BaseModel):
    id: Optional[int] = None
    nombre: str = Field(..., min_length=2, max_length=100, description="Nombre del cliente")
    telefono: str = Field(default="", max_length=50)
    direccion: str = Field(default="", max_length=300)
    zona: str = Field(default="", max_length=100)
    lista_precio_id: Optional[int] = None
    
    @field_validator('nombre')
    @classmethod
    def nombre_must_have_letter(cls, v: str) -> str:
        if not any(c.isalpha() for c in v):
            raise ValueError('El nombre debe contener al menos una letra')
        return v.strip()


class ClienteRef(BaseModel):
    id: Optional[int] = None
    nombre: Optional[str] = None
    telefono: Optional[str] = None
    direccion: Optional[str] = None
    zona: Optional[str] = None
    lista_precio_id: Optional[int] = None

class CategoriaCreate(BaseModel):
    """Modelo para crear/actualizar categorías con validación"""
    nombre: str = Field(..., min_length=1, max_length=100, description="Nombre de la categoría")
    descripcion: Optional[str] = Field(default="", max_length=500)
    color: str = Field(default="#6366f1", pattern="^#[0-9a-fA-F]{6}$", description="Color en formato hex")
    orden: int = Field(default=0, ge=0, description="Orden de visualización")
    activa: bool = Field(default=True)

class Producto(BaseModel):
    id: Optional[int] = None
    nombre: Optional[str] = Field(None, min_length=2, max_length=100)
    precio: Optional[float] = Field(None, gt=0, description="Precio debe ser mayor a 0")
    imagen_url: Optional[str] = None
    stock: Optional[float] = Field(default=0, ge=0, description="Stock no puede ser negativo")
    stock_minimo: Optional[float] = Field(default=10, ge=0)
    stock_tipo: Optional[str] = Field(default="unidad", pattern="^(unidad|caja|gancho|tira)$")
    categoria_id: Optional[int] = Field(default=None, description="ID de categoría (opcional)")

    @field_validator('stock', 'stock_minimo')
    @classmethod
    def validar_multiplo_medio(cls, v: Optional[float]) -> Optional[float]:
        if v is None:
            return v
        # Redondear a múltiplo de 0.5
        redondeado = round(v * 2) / 2
        if abs(v - redondeado) > 0.01:  # Tolerancia para floats
            raise ValueError("El valor debe ser múltiplo de 0.5 (ej: 0.5, 1, 1.5, 2, etc)")
        return redondeado

class ProductoConCantidad(Producto):
    cantidad: float
    tipo: str

    @field_validator('cantidad')
    @classmethod
    def validar_cantidad(cls, v: float) -> float:
        if v <= 0:
            raise ValueError("La cantidad debe ser mayor a 0")
        # Redondear a múltiplo de 0.5
        redondeado = round(v * 2) / 2
        if abs(v - redondeado) > 0.01:  # Tolerancia para floats
            raise ValueError("La cantidad debe ser múltiplo de 0.5 (ej: 0.5, 1, 1.5, 2, etc)")
        return redondeado

    @field_validator('tipo')
    @classmethod
    def validar_tipo(cls, v: str) -> str:
        if v not in ['unidad', 'caja', 'gancho', 'tira']:
            raise ValueError("El tipo debe ser 'unidad', 'caja', 'gancho' o 'tira'")
        return v

class Pedido(BaseModel):
    id: Optional[int] = None
    cliente: ClienteRef
    productos: List[ProductoConCantidad]
    fecha: Optional[str] = None
    pdf_generado: bool = False

class GenerarPDFsRequest(BaseModel):
    pedido_ids: List[int]

class PedidoItem(BaseModel):
    producto_id: int
    cantidad: float
    tipo: Optional[str] = "unidad"

    @field_validator('cantidad')
    @classmethod
    def validar_cantidad(cls, v: float) -> float:
        if v <= 0:
            raise ValueError("La cantidad debe ser mayor a 0")
        # Redondear a múltiplo de 0.5
        redondeado = round(v * 2) / 2
        if abs(v - redondeado) > 0.01:  # Tolerancia para floats
            raise ValueError("La cantidad debe ser múltiplo de 0.5 (ej: 0.5, 1, 1.5, 2, etc)")
        return redondeado

    @field_validator('tipo')
    @classmethod
    def validar_tipo(cls, v: str) -> str:
        if v not in ['unidad', 'caja', 'gancho', 'tira']:
            raise ValueError("El tipo debe ser 'unidad', 'caja', 'gancho' o 'tira'")
        return v

# === Health Check ===
def _health_check_impl():
    """Shared health check logic - returns health status dict"""
    try:
        # Test database connection with minimal query
        con = db.conectar()
        cur = con.cursor()
        cur.execute("SELECT 1")
        con.close()
        db_status = "ok"
    except Exception as e:
        db_status = f"error: {str(e)}"
    
    return {
        "status": "healthy" if db_status == "ok" else "degraded",
        "database": db_status,
        "version": API_VERSION,
        "timestamp": datetime.now(timezone.utc).isoformat()
    }

@app.api_route("/health", methods=["GET", "HEAD"], include_in_schema=False)
def health_check():
    """Health check endpoint for monitoring (supports GET and HEAD)"""
    return _health_check_impl()


@app.api_route("/healthz", methods=["GET", "HEAD"], include_in_schema=False)
def health_check_k8s():
    """Health check endpoint for Kubernetes/Render compatibility (alias)"""
    return _health_check_impl()


# Root endpoint (optional): return basic service info to avoid noisy 404s on /
@app.api_route("/", methods=["GET", "HEAD"], include_in_schema=False)
def root():
    return {"status": "ok", "service": "Chorizaurio API", "version": API_VERSION}

# === Autenticación ===
def validate_password_strength(password: str) -> tuple[bool, str]:
    """Valida la fortaleza del password. Retorna (is_valid, message)"""
    if len(password) < 8:
        return False, "La contraseña debe tener al menos 8 caracteres"
    if len(password) > 72:
        return False, "La contraseña no puede tener más de 72 caracteres"
    # Require at least one letter
    if not re.search(r'[A-Za-z]', password):
        return False, "La contraseña debe contener al menos una letra"
    # Require at least one number
    if not re.search(r'\d', password):
        return False, "La contraseña debe contener al menos un número"
    # Common passwords blocklist
    common_passwords = [
        '123456', '12345678', '123456789', 'password', 'admin', 'usuario',
        'qwerty', 'password1', 'admin123', '12341234', 'password123',
        'letmein', 'welcome', 'monkey', 'dragon', 'master', 'qwerty123',
        'iloveyou', 'trustno1', 'sunshine', '1234567890', 'princess',
        'football', 'baseball', 'abc123', 'login', 'passw0rd', 'shadow',
        'michael', 'superman', 'qwertyuiop', 'asdfghjkl', 'zxcvbnm'
    ]
    if password.lower() in common_passwords:
        return False, "Contraseña muy común. Elige una más segura"
    return True, ""


@app.post("/register")
@limiter.limit("5/minute")
def register(request: Request, form_data: OAuth2PasswordRequestForm = Depends(), rol: str = Form("usuario")):
    # Validar fortaleza del password
    is_valid, msg = validate_password_strength(form_data.password)
    if not is_valid:
        raise HTTPException(status_code=400, detail=msg)
    
    ctx = get_request_context(request)
    password_hash = hash_password(form_data.password)
    if db.add_usuario(form_data.username, password_hash, rol):
        logger.info(f"USER_REGISTERED: username={form_data.username} rol={rol} ip={ctx['ip']}")
        db.audit_log(form_data.username, "REGISTER", "usuarios", 
                     datos_despues={"username": form_data.username, "rol": rol},
                     ip_address=ctx["ip"], user_agent=ctx["user_agent"])
        return {"ok": True}
    raise HTTPException(status_code=400, detail="El usuario ya está registrado. Intenta con otro nombre.")

@app.post("/login")
@limiter.limit("5/minute")  # Strict rate limit for login attempts
def login(request: Request, username: str = Form(...), password: str = Form(...)):
    ctx = get_request_context(request)
    user = db.get_usuario(username)
    if not user or not verify_password(password, user["password_hash"]):
        logger.warning(f"LOGIN_FAILED: user={username} ip={ctx['ip']}")
        db.audit_log(username, "LOGIN_FAILED", "usuarios", ip_address=ctx["ip"], user_agent=ctx["user_agent"])
        raise HTTPException(status_code=401, detail="Usuario o contraseña incorrectos.")
    if not user["activo"]:
        logger.warning(f"LOGIN_INACTIVE: user={username} ip={ctx['ip']}")
        raise HTTPException(status_code=403, detail="Tu cuenta está inactiva. Contacta al administrador.")
    logger.info(f"LOGIN_SUCCESS: user={username} rol={user['rol']} ip={ctx['ip']}")
    db.audit_log(username, "LOGIN", "usuarios", ip_address=ctx["ip"], user_agent=ctx["user_agent"])
    token = create_access_token(data={"sub": username, "rol": user["rol"], "activo": user["activo"]})
    return {"access_token": token, "token_type": "bearer"}


@app.post("/logout")
@limiter.limit("30/minute")
def logout(request: Request, token: str = Depends(oauth2_scheme)):
    """Logout by revoking the current token"""
    ctx = get_request_context(request)
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username = payload.get("sub")
        jti = payload.get("jti")
        exp_iso = payload.get("exp_iso", "")
        
        if jti:
            db.revoke_token(jti, exp_iso, username)
            logger.info(f"LOGOUT: user={username} ip={ctx['ip']}")
            db.audit_log(username, "LOGOUT", "usuarios", ip_address=ctx["ip"], user_agent=ctx["user_agent"])
            return {"message": "Sesión cerrada correctamente"}
        else:
            # Token without jti (old token before this feature)
            logger.warning(f"LOGOUT_OLD_TOKEN: user={username} - token lacks jti")
            return {"message": "Sesión cerrada (token antiguo)"}
    except JWTError:
        raise HTTPException(status_code=401, detail="Token inválido")


@app.post("/refresh")
@limiter.limit("30/minute")
def refresh_token(request: Request, user=Depends(get_current_user)):
    """Renueva el token JWT si el actual es válido"""
    ctx = get_request_context(request)
    # Verificar que el usuario sigue activo en la base de datos
    db_user = db.get_usuario(user["username"])
    if not db_user or not db_user["activo"]:
        raise HTTPException(status_code=401, detail="Usuario inactivo o no encontrado")
    
    # Generar nuevo token con datos actualizados
    new_token = create_access_token(data={
        "sub": db_user["username"], 
        "rol": db_user["rol"], 
        "activo": db_user["activo"]
    })
    logger.info(f"TOKEN_REFRESH: user={user['username']} ip={ctx['ip']}")
    return {"access_token": new_token, "token_type": "bearer"}


@app.post("/cambiar-password")
@limiter.limit("5/minute")
def cambiar_password(
    request: Request, 
    password_actual: str = Form(...),
    password_nuevo: str = Form(...),
    user=Depends(get_current_user)
):
    """Permite al usuario cambiar su propia contraseña"""
    ctx = get_request_context(request)
    
    # Verificar password actual
    db_user = db.get_usuario(user["username"])
    if not db_user or not verify_password(password_actual, db_user["password_hash"]):
        logger.warning(f"PASSWORD_CHANGE_FAILED: user={user['username']} reason=wrong_current_password ip={ctx['ip']}")
        db.audit_log(user["username"], "PASSWORD_CHANGE_FAILED", "usuarios", 
                     ip_address=ctx["ip"], user_agent=ctx["user_agent"])
        raise HTTPException(status_code=400, detail="La contraseña actual es incorrecta")
    
    # Validar nuevo password
    is_valid, msg = validate_password_strength(password_nuevo)
    if not is_valid:
        raise HTTPException(status_code=400, detail=msg)
    
    # No permitir reusar la misma contraseña
    if verify_password(password_nuevo, db_user["password_hash"]):
        raise HTTPException(status_code=400, detail="La nueva contraseña debe ser diferente a la actual")
    
    # Actualizar password
    new_hash = hash_password(password_nuevo)
    db.update_usuario_password(user["username"], new_hash)
    
    logger.info(f"PASSWORD_CHANGED: user={user['username']} ip={ctx['ip']}")
    db.audit_log(user["username"], "PASSWORD_CHANGE", "usuarios", 
                 ip_address=ctx["ip"], user_agent=ctx["user_agent"])
    
    return {"message": "Contraseña actualizada correctamente"}


# === Endpoints protegidos ===
@app.get("/clientes")
def get_clientes(
    page: Optional[int] = Query(default=None, ge=1, description="Número de página"), 
    limit: int = Query(default=50, ge=1, le=200, description="Resultados por página (máx 200)"), 
    search: Optional[str] = Query(default=None, max_length=100, description="Búsqueda por nombre/teléfono"),
    user=Depends(get_current_user)
):
    """Lista clientes con paginación y búsqueda opcional."""
    return db.get_clientes(page=page, limit=limit, search=search)


@app.get("/clientes/{cliente_id}")
def get_cliente(cliente_id: int, user=Depends(get_current_user)):
    """Obtiene un cliente por su ID."""
    cliente = db.get_cliente_by_id(cliente_id)
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    return cliente


@app.post("/clientes")
@limiter.limit("30/minute")
def add_cliente(request: Request, cliente: Cliente, user=Depends(get_current_user)):
    ctx = get_request_context(request)
    res = db.add_cliente(cliente.model_dump())
    if isinstance(res, dict) and res.get("error"):
        raise HTTPException(status_code=400, detail=res)
    logger.info(f"CLIENTE_CREATED: id={res.get('id')} by={user['username']}")
    db.audit_log(user["username"], "CREATE", "clientes", registro_id=res.get("id"), 
                 datos_despues=res, ip_address=ctx["ip"], user_agent=ctx["user_agent"])
    return res

@app.put("/clientes/{cliente_id}")
@limiter.limit("30/minute")
def update_cliente(request: Request, cliente_id: int, cliente: Cliente, user=Depends(get_current_user)):
    ctx = get_request_context(request)
    old = db.get_cliente_by_id(cliente_id)
    res = db.update_cliente(cliente_id, cliente.model_dump())
    if isinstance(res, dict) and res.get("error"):
        raise HTTPException(status_code=400, detail=res)
    logger.info(f"CLIENTE_UPDATED: id={cliente_id} by={user['username']}")
    db.audit_log(user["username"], "UPDATE", "clientes", registro_id=cliente_id,
                 datos_antes=old, datos_despues=res, ip_address=ctx["ip"], user_agent=ctx["user_agent"])
    return res

@app.delete("/clientes/{cliente_id}")
@limiter.limit("30/minute")
def delete_cliente(cliente_id: int, request: Request, user=Depends(get_admin_user)):
    ctx = get_request_context(request)
    old = db.get_cliente_by_id(cliente_id)
    logger.info(f"CLIENTE_DELETED: id={cliente_id} by={user['username']}")
    db.audit_log(user["username"], "DELETE", "clientes", registro_id=cliente_id,
                 datos_antes=old, ip_address=ctx["ip"], user_agent=ctx["user_agent"])
    return db.delete_cliente(cliente_id)


@app.post("/clientes/bulk-delete")
@limiter.limit("10/minute")
def bulk_delete_clientes(request: Request, ids: List[int], user=Depends(get_admin_user)):
    """Elimina múltiples clientes a la vez"""
    ctx = get_request_context(request)
    if not ids:
        raise HTTPException(status_code=400, detail="No se proporcionaron IDs")
    if len(ids) > 50:
        raise HTTPException(status_code=400, detail="Máximo 50 clientes por operación")
    
    deleted = 0
    errors = []
    for cliente_id in ids:
        try:
            old = db.get_cliente_by_id(cliente_id)
            if old:
                db.delete_cliente(cliente_id)
                db.audit_log(user["username"], "DELETE", "clientes", registro_id=cliente_id,
                             datos_antes=old, ip_address=ctx["ip"], user_agent=ctx["user_agent"])
                deleted += 1
        except Exception as e:
            errors.append({"id": cliente_id, "error": str(e)})
    
    logger.info(f"BULK_DELETE_CLIENTES: deleted={deleted} by={user['username']}")
    return {"deleted": deleted, "errors": errors}


@app.get("/clientes/export/csv")
def export_clientes_csv(user=Depends(get_admin_user)):
    from fastapi.responses import Response
    logger.info(f"CSV export: clientes by {user['username']}")
    csv_content = db.export_clientes_csv()
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=clientes.csv"}
    )


# === Categorías ===
@app.get("/categorias")
def get_categorias(incluir_inactivas: bool = False, user=Depends(get_current_user)):
    """Obtiene todas las categorías."""
    return db.get_categorias(incluir_inactivas=incluir_inactivas)


@app.get("/categorias/{categoria_id}")
def get_categoria(categoria_id: int, user=Depends(get_current_user)):
    """Obtiene una categoría por ID."""
    categoria = db.get_categoria_by_id(categoria_id)
    if not categoria:
        raise HTTPException(status_code=404, detail="Categoría no encontrada")
    return categoria


@app.post("/categorias")
@limiter.limit("30/minute")
def add_categoria(request: Request, categoria: CategoriaCreate, user=Depends(get_admin_user)):
    """Crea una nueva categoría con validación."""
    res = db.add_categoria(categoria.model_dump())
    if isinstance(res, dict) and res.get("error"):
        raise HTTPException(status_code=400, detail=res["error"])
    db.audit_log(user["username"], "CREATE", "categorias", registro_id=res.get("id"), datos_despues=res)
    return res


@app.put("/categorias/{categoria_id}")
@limiter.limit("30/minute")
def update_categoria(request: Request, categoria_id: int, categoria: CategoriaCreate, user=Depends(get_admin_user)):
    """Actualiza una categoría con validación."""
    old = db.get_categoria_by_id(categoria_id)
    res = db.update_categoria(categoria_id, categoria.model_dump())
    if isinstance(res, dict) and res.get("error"):
        raise HTTPException(status_code=400, detail=res["error"])
    db.audit_log(user["username"], "UPDATE", "categorias", registro_id=categoria_id, datos_antes=old, datos_despues=res)
    return res


@app.delete("/categorias/{categoria_id}")
@limiter.limit("30/minute")
def delete_categoria(categoria_id: int, request: Request, user=Depends(get_admin_user)):
    """Elimina o desactiva una categoría."""
    ctx = get_request_context(request)
    old = db.get_categoria_by_id(categoria_id)
    db.audit_log(user["username"], "DELETE", "categorias", registro_id=categoria_id, datos_antes=old, ip_address=ctx["ip"], user_agent=ctx["user_agent"])
    return db.delete_categoria(categoria_id)


@app.get("/categorias/{categoria_id}/productos")
def get_productos_por_categoria(categoria_id: int, user=Depends(get_current_user)):
    """Obtiene productos de una categoría."""
    return db.get_productos_por_categoria(categoria_id)


# === Tags (Multi-etiqueta para productos) ===
@app.get("/tags")
def get_tags(tipo: Optional[str] = None, user=Depends(get_current_user)):
    """
    Obtiene todos los tags del sistema.
    - tipo: 'conservacion' | 'tipo' para filtrar por tipo de tag
    """
    return db.get_tags(tipo=tipo)


@app.get("/productos/{producto_id}/tags")
def get_producto_tags(producto_id: int, user=Depends(get_current_user)):
    """Obtiene los tags asignados a un producto."""
    producto = db.get_producto_by_id(producto_id)
    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    return db.get_producto_tags(producto_id)


@app.put("/productos/{producto_id}/tags")
@limiter.limit("60/minute")
def update_producto_tags(
    request: Request,
    producto_id: int,
    tag_ids: List[int],
    user=Depends(get_current_user)
):
    """Actualiza los tags de un producto (reemplaza todos los existentes)."""
    ctx = get_request_context(request)
    producto = db.get_producto_by_id(producto_id)
    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    
    old_tags = db.get_producto_tags(producto_id)
    result = db.update_producto_tags(producto_id, tag_ids)
    
    if result.get("error"):
        raise HTTPException(status_code=400, detail=result["error"])
    
    new_tags = db.get_producto_tags(producto_id)
    db.audit_log(
        user["username"], "UPDATE_TAGS", "productos_tags",
        registro_id=producto_id,
        datos_antes={"tags": old_tags},
        datos_despues={"tags": new_tags},
        ip_address=ctx["ip"], user_agent=ctx["user_agent"]
    )
    return {"producto_id": producto_id, "tags": new_tags}


@app.get("/tags/{tag_id}/productos")
def get_productos_por_tag(tag_id: int, user=Depends(get_current_user)):
    """Obtiene productos que tienen un tag específico."""
    return db.get_productos_por_tag(tag_id)


@app.get("/productos-con-tags")
def get_productos_with_tags(user=Depends(get_current_user)):
    """
    Obtiene todos los productos con sus tags incluidos en una sola consulta.
    Evita el problema N+1 de obtener tags individualmente.
    """
    return db.get_productos_with_tags()


@app.put("/productos/{producto_id}/categoria")
def asignar_categoria(producto_id: int, categoria_id: Optional[int] = None, user=Depends(get_current_user)):
    """Asigna una categoría a un producto."""
    db.audit_log(user["username"], "UPDATE", "productos", registro_id=producto_id, datos_despues={"categoria_id": categoria_id})
    return db.asignar_categoria_producto(producto_id, categoria_id)


# === Audit Log ===
@app.get("/admin/audit-logs")
def get_audit_logs(
    tabla: Optional[str] = None,
    usuario: Optional[str] = None,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    limit: int = 100,
    offset: int = 0,
    user=Depends(get_admin_user)
):
    """Obtiene registros del audit log."""
    # Validate tabla parameter against whitelist
    if tabla and tabla not in db.VALID_TABLES:
        raise HTTPException(status_code=400, detail=f"Tabla inválida: {tabla}")
    return db.get_audit_logs(
        tabla=tabla,
        usuario=usuario,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        limit=limit,
        offset=offset
    )


@app.get("/admin/audit-summary")
def get_audit_summary(user=Depends(get_admin_user)):
    """Resumen de actividad del audit log."""
    return db.get_audit_summary()


@app.get("/productos")
def get_productos(q: Optional[str] = None, sort: Optional[str] = None, categoria_id: Optional[int] = None, user=Depends(get_current_user)):
    return db.get_productos(search=q, sort=sort, categoria_id=categoria_id)

@app.get("/productos/export/csv")
def export_productos_csv(user=Depends(get_admin_user)):
    from fastapi.responses import Response
    logger.info(f"CSV export: productos by {user['username']}")
    csv_content = db.export_productos_csv()
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=productos.csv"}
    )

@app.get("/productos/export/xlsx")
def export_productos_xlsx(user=Depends(get_admin_user)):
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    
    logger.info(f"Excel export: productos by {user['username']}")
    productos = db.get_productos()
    categorias = {c['id']: c['nombre'] for c in db.get_categorias()}
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["ID", "Nombre", "Precio", "Stock", "Tipo Stock", "Stock Mínimo", "Categoría", "Imagen URL"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # Datos
    for row_num, p in enumerate(productos, 2):
        categoria_nombre = categorias.get(p.get('categoria_id'), "Sin categoría")
        row_data = [
            p.get('id', ''),
            p.get('nombre', ''),
            p.get('precio', 0),
            p.get('stock', 0),
            p.get('stock_tipo', 'unidad'),
            p.get('stock_minimo', 10),
            categoria_nombre,
            p.get('imagen_url', '')
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            if col == 3:  # Precio
                cell.number_format = '"$"#,##0.00'
    
    # Ajustar anchos de columna
    column_widths = [8, 35, 12, 10, 12, 14, 20, 50]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=productos.xlsx"}
    )

@app.put("/productos/{producto_id}/stock")
def update_stock(
    request: Request,
    producto_id: int, 
    cantidad: int, 
    operacion: str = "sumar",
    user=Depends(get_current_user)
):
    """Actualizar stock de un producto. operacion: 'sumar' o 'restar'"""
    ctx = get_request_context(request)
    if operacion not in ["sumar", "restar"]:
        raise HTTPException(status_code=400, detail="Operación inválida. Usa 'sumar' o 'restar'.")
    result = db.update_stock(producto_id, cantidad, operacion)
    logger.info(f"STOCK_UPDATED: producto={producto_id} {operacion}={cantidad} by={user['username']}")
    db.audit_log(user["username"], f"STOCK_{operacion.upper()}", "productos", registro_id=producto_id,
                 datos_antes={"stock": result.get("stock_anterior")},
                 datos_despues={"stock": result.get("stock_nuevo")},
                 ip_address=ctx["ip"], user_agent=ctx["user_agent"])
    return result


@app.get("/productos/{producto_id}")
def get_producto(producto_id: int, user=Depends(get_current_user)):
    producto = db.get_producto_by_id(producto_id)
    if not producto:
        raise HTTPException(status_code=404, detail='Producto no encontrado. Verifica el ID.')
    return producto

@app.post("/productos")
@limiter.limit("60/minute")
def add_producto(request: Request, producto: Producto, user=Depends(get_current_user)):
    ctx = get_request_context(request)
    if producto.imagen_url and not _is_valid_image_url(producto.imagen_url):
        raise HTTPException(status_code=400, detail="imagen_url inválida")
    res = db.add_producto(producto.model_dump())
    if isinstance(res, dict) and res.get("error"):
        raise HTTPException(status_code=400, detail=res)
    logger.info(f"PRODUCTO_CREATED: id={res.get('id')} nombre={producto.nombre} by={user['username']}")
    db.audit_log(user["username"], "CREATE", "productos", registro_id=res.get("id"),
                 datos_despues=res, ip_address=ctx["ip"], user_agent=ctx["user_agent"])
    return res


@app.put("/productos/{producto_id}")
@limiter.limit("60/minute")
def put_producto(request: Request, producto_id: int, producto: Producto, user=Depends(get_current_user)):
    ctx = get_request_context(request)
    old = db.get_producto_by_id(producto_id)
    if producto.imagen_url and not _is_valid_image_url(producto.imagen_url):
        raise HTTPException(status_code=400, detail="imagen_url inválida")
    res = db.update_producto(producto_id, producto.model_dump())
    if isinstance(res, dict) and res.get("error"):
        raise HTTPException(status_code=400, detail=res)
    logger.info(f"PRODUCTO_UPDATED: id={producto_id} by={user['username']}")
    db.audit_log(user["username"], "UPDATE", "productos", registro_id=producto_id,
                 datos_antes=old, datos_despues=res, ip_address=ctx["ip"], user_agent=ctx["user_agent"])
    return res


@app.delete("/productos/{producto_id}")
@limiter.limit("30/minute")
def delete_producto(request: Request, producto_id: int, user=Depends(get_current_user)):
    """Elimina un producto si no tiene pedidos asociados"""
    ctx = get_request_context(request)
    if user.get("rol") != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores pueden eliminar productos")
    old = db.get_producto_by_id(producto_id)
    res = db.delete_producto(producto_id)
    if res.get("status") == "not_found":
        raise HTTPException(status_code=404, detail=res["message"])
    if res.get("status") == "error":
        raise HTTPException(status_code=400, detail=res["message"])
    logger.info(f"PRODUCTO_DELETED: id={producto_id} by={user['username']}")
    db.audit_log(user["username"], "DELETE", "productos", registro_id=producto_id,
                 datos_antes=old, ip_address=ctx["ip"], user_agent=ctx["user_agent"])
    return res


@app.post("/productos/bulk-delete")
@limiter.limit("10/minute")
def bulk_delete_productos(request: Request, ids: List[int], user=Depends(get_current_user)):
    """Elimina múltiples productos a la vez (solo admin)"""
    ctx = get_request_context(request)
    if user.get("rol") != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores pueden eliminar productos")
    if not ids:
        raise HTTPException(status_code=400, detail="No se proporcionaron IDs")
    if len(ids) > 50:
        raise HTTPException(status_code=400, detail="Máximo 50 productos por operación")
    
    deleted = 0
    errors = []
    for producto_id in ids:
        try:
            old = db.get_producto_by_id(producto_id)
            if old:
                res = db.delete_producto(producto_id)
                if res.get("status") == "ok":
                    db.audit_log(user["username"], "DELETE", "productos", registro_id=producto_id,
                                 datos_antes=old, ip_address=ctx["ip"], user_agent=ctx["user_agent"])
                    deleted += 1
                else:
                    errors.append({"id": producto_id, "error": res.get("message", "Error desconocido")})
        except Exception as e:
            errors.append({"id": producto_id, "error": str(e)})
    
    logger.info(f"BULK_DELETE_PRODUCTOS: deleted={deleted} by={user['username']}")
    return {"deleted": deleted, "errors": errors}


@app.post('/upload')
@limiter.limit("20/minute")
def upload_file(request: Request, file: UploadFile = File(...), user=Depends(get_current_user)):
    """Recibe multipart file y lo guarda en /data/uploads, retorna URL relativa."""
    # Security: Validate file extension
    filename = os.path.basename(file.filename or "")
    ext = os.path.splitext(filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"Tipo de archivo no permitido. Extensiones válidas: {', '.join(ALLOWED_EXTENSIONS)}")
    
    # Security: Validate MIME type
    content_type = file.content_type or ""
    if content_type not in ALLOWED_MIME_TYPES:
        raise HTTPException(status_code=400, detail=f"Tipo MIME no permitido: {content_type}")
    
    # Security: Check file size (read in chunks)
    safe_name = f"{uuid4().hex}{ext}"
    dest = os.path.join(UPLOAD_DIR, safe_name)
    total_size = 0
    
    try:
        with open(dest, 'wb') as out:
            while chunk := file.file.read(8192):  # 8KB chunks
                total_size += len(chunk)
                if total_size > MAX_UPLOAD_SIZE:
                    out.close()
                    os.remove(dest)
                    raise HTTPException(status_code=400, detail=f"Archivo demasiado grande. Máximo: {MAX_UPLOAD_SIZE // (1024*1024)}MB")
                out.write(chunk)
    finally:
        file.file.close()
    
    logger.info(f"File uploaded: {safe_name} ({total_size} bytes) by {user.get('username')}")
    return {"url": f"/media/uploads/{safe_name}"}

@app.get("/pedidos")
def get_pedidos(user=Depends(get_current_user)):
    return db.get_pedidos()

@app.get("/pedidos/export/csv")
def export_pedidos_csv(
    desde: Optional[str] = None, 
    hasta: Optional[str] = None,
    user=Depends(get_admin_user)
):
    from fastapi.responses import Response
    logger.info(f"CSV export: pedidos by {user['username']}")
    csv_content = db.export_pedidos_csv(desde=desde, hasta=hasta)
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=pedidos.csv"}
    )

@app.post("/pedidos/verificar-stock")
def verificar_stock_pedido(pedido: Pedido, user=Depends(get_current_user)):
    """Verifica si hay stock suficiente para los productos del pedido"""
    productos = [{"id": p.id, "cantidad": p.cantidad} for p in pedido.productos]
    errores = db.verificar_stock_pedido(productos)
    if errores:
        return {"ok": False, "errores": errores}
    return {"ok": True, "errores": []}

@app.post("/pedidos")
def add_pedido(pedido: Pedido, request: Request, user=Depends(get_current_user)):
    # Require client with valid ID
    if not pedido.cliente or not pedido.cliente.id:
        raise HTTPException(
            status_code=400,
            detail="Debes seleccionar un cliente para crear el pedido."
        )
    
    # Verify client exists
    clientes = db.get_clientes()
    if isinstance(clientes, dict) and clientes.get('data'):
        clientes = clientes['data']
    cliente_existe = any(c.get('id') == pedido.cliente.id for c in clientes)
    if not cliente_existe:
        raise HTTPException(
            status_code=400,
            detail="El cliente seleccionado no existe."
        )
    
    # Verificar stock antes de crear pedido
    productos = [{"id": p.id, "cantidad": p.cantidad} for p in pedido.productos]
    errores = db.verificar_stock_pedido(productos)
    if errores:
        raise HTTPException(
            status_code=400, 
            detail={"message": "Stock insuficiente para algunos productos. Ver detalles abajo.", "errores": errores}
        )
    
    # Detectar dispositivo y user agent
    dispositivo, user_agent = _detect_device(request)
    
    # Pasar username del usuario que crea el pedido para tracking
    result = db.add_pedido(
        pedido.model_dump(), 
        creado_por=user.get('username'),
        dispositivo=dispositivo,
        user_agent=user_agent
    )
    
    # Registrar en historial
    db.add_historial(result.get('id'), 'CREADO', user.get('username'), f'Pedido creado desde {dispositivo}')
    
    return result

@app.delete("/pedidos/{pedido_id}")
@limiter.limit("30/minute")
def delete_pedido(pedido_id: int, request: Request, user=Depends(get_admin_user)):
    logger.info(f"DELETE /pedidos/{pedido_id} by {user['username']}")
    # El historial se elimina automáticamente en db.delete_pedido (cascade)
    return db.delete_pedido(pedido_id)

@app.patch("/pedidos/{pedido_id}")
def update_pedido_estado(pedido_id: int, user=Depends(get_current_user)):
    return db.update_pedido_estado(pedido_id, 1)

@app.put("/pedidos/{pedido_id}/cliente")
def update_pedido_cliente(pedido_id: int, cliente_id: int, user=Depends(get_current_user)):
    """Actualizar el cliente de un pedido existente"""
    result = db.update_pedido_cliente(pedido_id, cliente_id)
    # Registrar edición
    db.update_pedido_ultima_edicion(pedido_id, user.get('username'))
    db.add_historial(pedido_id, 'CLIENTE_CAMBIADO', user.get('username'), f'Cliente cambiado a ID {cliente_id}')
    return result


class NotasUpdate(BaseModel):
    notas: str = ""


@app.put("/pedidos/{pedido_id}/notas")
def update_pedido_notas(pedido_id: int, data: NotasUpdate, user=Depends(get_current_user)):
    """Actualizar las notas de un pedido"""
    result = db.update_pedido_notas(pedido_id, data.notas)
    db.update_pedido_ultima_edicion(pedido_id, user.get('username'))
    db.add_historial(pedido_id, 'NOTAS_ACTUALIZADAS', user.get('username'), f'Notas actualizadas')
    return result

@app.post("/pedidos/preview_stock")
def preview_stock_for_pdfs(request: GenerarPDFsRequest, user=Depends(get_current_user)):
    """
    Preview stock changes before generating PDFs. Returns warnings if stock would go to zero.
    """
    todos_pedidos = db.get_pedidos()
    pedidos_a_generar = [p for p in todos_pedidos if p.get('id') in request.pedido_ids]
    
    # Get all productos to check current stock
    productos_dict = {p['id']: p for p in db.get_productos()}
    stock_changes = {}
    warnings = []
    
    for pedido in pedidos_a_generar:
        for prod in pedido.get('productos', []):
            pid = prod.get('id')
            cantidad = prod.get('cantidad', 0)
            if pid not in stock_changes:
                stock_changes[pid] = {
                    'nombre': prod.get('nombre'),
                    'stock_actual': productos_dict.get(pid, {}).get('stock', 0),
                    'cantidad_a_restar': 0
                }
            stock_changes[pid]['cantidad_a_restar'] += cantidad
    
    for pid, info in stock_changes.items():
        nuevo_stock = info['stock_actual'] - info['cantidad_a_restar']
        info['stock_nuevo'] = max(0, nuevo_stock)
        if nuevo_stock <= 0:
            warnings.append({
                'producto': info['nombre'],
                'stock_actual': info['stock_actual'],
                'cantidad_pedida': info['cantidad_a_restar'],
                'mensaje': f"⚠️ {info['nombre']} quedará en 0 (actual: {info['stock_actual']}, pedido: {info['cantidad_a_restar']})"
            })
    
    return {
        "ok": True,
        "pedidos_count": len(pedidos_a_generar),
        "stock_changes": list(stock_changes.values()),
        "warnings": warnings
    }

@app.post("/pedidos/generar_pdfs")
@limiter.limit("10/minute")
def generar_pdfs(request: Request, body: GenerarPDFsRequest, user=Depends(get_current_user)):
    """
    Genera PDF para los pedidos seleccionados y resta stock de productos.
    Al generar PDF, se marca como 'generado' y se resta stock.
    Returns the PDF as binary response.
    
    Uses atomic transaction pattern: if any step fails, changes are rolled back.
    """
    ctx = get_request_context(request)
    
    # Obtener pedidos
    todos_pedidos = db.get_pedidos()
    pedidos_a_generar = [p for p in todos_pedidos if p.get('id') in body.pedido_ids]
    
    if not pedidos_a_generar:
        raise HTTPException(status_code=400, detail="No hay pedidos para generar")
    
    # Username del usuario que genera el PDF
    generado_por = user.get('username')
    
    # Track what we've done for potential rollback
    processed_pedidos = []
    stock_changes = []
    
    try:
        # Restar stock de forma atómica para cada pedido
        for pedido in pedidos_a_generar:
            productos = pedido.get('productos', [])
            # Usar transacción atómica para actualizar stock
            productos_stock = [{"id": p.get('id'), "cantidad": p.get('cantidad', 1)} for p in productos]
            result = db.batch_update_stock_atomic(productos_stock, operacion='restar')
            
            if result.get('error'):
                # Rollback previously processed stock changes
                for change in stock_changes:
                    db.batch_update_stock_atomic(change, operacion='sumar')
                raise HTTPException(status_code=400, detail=f"Error actualizando stock: {result['error']}")
            
            stock_changes.append(productos_stock)
            
            # Marcar pedido como generado con tracking de usuario y fecha
            db.update_pedido_estado(pedido.get('id'), 1, generado_por=generado_por)
            
            # Registrar en historial
            db.add_historial(pedido.get('id'), 'PDF_GENERADO', generado_por, f'PDF generado ({len(productos)} productos)')
            processed_pedidos.append(pedido.get('id'))
        
        # Log PDF generation
        pedido_ids = [p.get('id') for p in pedidos_a_generar]
        logger.info(f"PDF_GENERATED: pedidos={pedido_ids} count={len(pedidos_a_generar)} by={generado_por}")
        db.audit_log(generado_por, "PDF_GENERATE", "pedidos", 
                     datos_despues={"pedido_ids": pedido_ids, "count": len(pedidos_a_generar)},
                     ip_address=ctx["ip"], user_agent=ctx["user_agent"])
        
        # Generate PDF
        clientes = db.get_clientes()
        fecha_generacion = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        pdf_bytes = pdf_utils.generar_pdf_multiple(pedidos_a_generar, clientes, fecha_generacion)
        
        # Return PDF as binary response
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=pedidos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            }
        )
    except HTTPException:
        raise  # Re-raise HTTP exceptions as-is
    except Exception as e:
        # Rollback all stock changes on unexpected error
        logger.error(f"PDF_GENERATION_FAILED: error={str(e)}, rolling back {len(stock_changes)} stock changes")
        for change in stock_changes:
            try:
                db.batch_update_stock_atomic(change, operacion='sumar')
            except Exception as rollback_err:
                logger.error(f"ROLLBACK_FAILED: {rollback_err}")
        raise HTTPException(status_code=500, detail=f"Error generando PDF: {str(e)}")


# === Edición de ítems de pedidos pendientes ===
@app.post("/pedidos/{pedido_id}/items")
def add_pedido_item(pedido_id: int, item: PedidoItem, user=Depends(get_current_user)):
    res = db.add_pedido_item(pedido_id, item.producto_id, item.cantidad, item.tipo or "unidad")
    if isinstance(res, dict) and res.get("error"):
        raise HTTPException(status_code=400, detail=res)
    # Registrar edición
    db.update_pedido_ultima_edicion(pedido_id, user.get('username'))
    db.add_historial(pedido_id, 'PRODUCTO_AGREGADO', user.get('username'), f'Producto ID {item.producto_id} agregado (cant: {item.cantidad})')
    return res

@app.put("/pedidos/{pedido_id}/items/{producto_id}")
def update_pedido_item(pedido_id: int, producto_id: int, item: PedidoItem, user=Depends(get_current_user)):
    res = db.update_pedido_item(pedido_id, producto_id, item.cantidad, item.tipo)
    if isinstance(res, dict) and res.get("error"):
        raise HTTPException(status_code=400, detail=res)
    # Registrar edición
    db.update_pedido_ultima_edicion(pedido_id, user.get('username'))
    db.add_historial(pedido_id, 'PRODUCTO_EDITADO', user.get('username'), f'Producto ID {producto_id} editado (cant: {item.cantidad})')
    return res

@app.delete("/pedidos/{pedido_id}/items/{producto_id}")
@limiter.limit("30/minute")
def delete_pedido_item(pedido_id: int, producto_id: int, request: Request, user=Depends(get_admin_user)):
    logger.info(f"DELETE /pedidos/{pedido_id}/items/{producto_id} by {user['username']}")
    res = db.delete_pedido_item(pedido_id, producto_id)
    if isinstance(res, dict) and res.get("error"):
        raise HTTPException(status_code=400, detail=res)
    # Registrar edición
    db.update_pedido_ultima_edicion(pedido_id, user.get('username'))
    db.add_historial(pedido_id, 'PRODUCTO_ELIMINADO', user.get('username'), f'Producto ID {producto_id} eliminado')
    return res


# === Dashboard endpoints (Semana 1) ===
import dashboard as dash

@app.get("/dashboard/metrics")
def get_dashboard_metrics(user=Depends(get_current_user)):
    """Get main KPI metrics for dashboard"""
    return dash.get_dashboard_metrics()

@app.get("/dashboard/pedidos_por_dia")
def get_pedidos_por_dia(dias: int = 30, user=Depends(get_current_user)):
    """Get orders per day chart data"""
    return dash.get_pedidos_por_dia(dias)

@app.get("/dashboard/alertas")
def get_alertas(user=Depends(get_current_user)):
    """Get system alerts (stock bajo, etc)"""
    return dash.get_alertas()


# === Ofertas endpoints ===
@app.get("/ofertas/activas")
def get_ofertas_activas(user=Depends(get_current_user)):
    """Get active offers with their product IDs"""
    return db.get_ofertas_activas()

@app.get("/ofertas")
def get_all_ofertas(user=Depends(get_current_user)):
    """Get all offers (active and inactive)"""
    return db.get_todas_ofertas()

@app.post("/ofertas")
def create_oferta(
    request: Request,
    titulo: str = Form(...),
    descripcion: str = Form(""),
    desde: str = Form(...),
    hasta: str = Form(...),
    productos: str = Form(...),
    descuento_porcentaje: float = Form(10),
    user=Depends(get_admin_user)
):
    """Create a new offer with products and quantities"""
    import json
    ctx = get_request_context(request)
    productos_list = json.loads(productos)
    res = db.crear_oferta(titulo, descripcion, desde, hasta, productos_list, descuento_porcentaje)
    logger.info(f"OFERTA_CREATED: id={res.get('id')} titulo={titulo} by={user['username']}")
    db.audit_log(user["username"], "CREATE", "ofertas", registro_id=res.get("id"),
                 datos_despues={"titulo": titulo, "desde": desde, "hasta": hasta, "descuento": descuento_porcentaje},
                 ip_address=ctx["ip"], user_agent=ctx["user_agent"])
    return res

@app.put("/ofertas/{oferta_id}")
def update_oferta(
    request: Request,
    oferta_id: int,
    titulo: str = Form(...),
    descripcion: str = Form(""),
    desde: str = Form(...),
    hasta: str = Form(...),
    productos: str = Form(...),
    descuento_porcentaje: float = Form(10),
    user=Depends(get_admin_user)  # Cambiado a admin_user por seguridad
):
    """Update an existing offer with products and quantities"""
    import json
    ctx = get_request_context(request)
    productos_list = json.loads(productos)
    res = db.actualizar_oferta(oferta_id, titulo, descripcion, desde, hasta, productos_list, descuento_porcentaje)
    logger.info(f"OFERTA_UPDATED: id={oferta_id} by={user['username']}")
    db.audit_log(user["username"], "UPDATE", "ofertas", registro_id=oferta_id,
                 datos_despues={"titulo": titulo, "desde": desde, "hasta": hasta, "descuento": descuento_porcentaje},
                 ip_address=ctx["ip"], user_agent=ctx["user_agent"])
    return res

@app.delete("/ofertas/{oferta_id}")
@limiter.limit("30/minute")
def delete_oferta(request: Request, oferta_id: int, user=Depends(get_admin_user)):
    """Delete an offer"""
    ctx = get_request_context(request)
    logger.info(f"OFERTA_DELETED: id={oferta_id} by={user['username']}")
    db.audit_log(user["username"], "DELETE", "ofertas", registro_id=oferta_id,
                 ip_address=ctx["ip"], user_agent=ctx["user_agent"])
    db.eliminar_oferta(oferta_id)
    return {"message": "Oferta eliminada"}

@app.post("/ofertas/{oferta_id}/toggle")
def toggle_oferta(request: Request, oferta_id: int, user=Depends(get_admin_user)):
    """Toggle offer active status"""
    ctx = get_request_context(request)
    res = db.toggle_oferta(oferta_id)
    logger.info(f"OFERTA_TOGGLED: id={oferta_id} activa={res.get('activa')} by={user['username']}")
    db.audit_log(user["username"], "TOGGLE", "ofertas", registro_id=oferta_id,
                 datos_despues={"activa": res.get("activa")},
                 ip_address=ctx["ip"], user_agent=ctx["user_agent"])
    return res


# === LISTAS DE PRECIOS ===
@app.get("/listas-precios")
def get_listas_precios(user=Depends(get_current_user)):
    """Obtiene todas las listas de precios"""
    return db.get_listas_precios()


@app.get("/listas-precios/{lista_id}")
def get_lista_precio(lista_id: int, user=Depends(get_current_user)):
    """Obtiene una lista de precios con sus precios especiales"""
    lista = db.get_lista_precio(lista_id)
    if not lista:
        raise HTTPException(404, "Lista no encontrada")
    return lista


class ListaPrecioCreate(BaseModel):
    nombre: str
    descripcion: Optional[str] = ""
    multiplicador: float = 1.0


@app.post("/listas-precios")
def create_lista_precio(data: ListaPrecioCreate, user=Depends(get_admin_user)):
    """Crea una nueva lista de precios"""
    try:
        result = db.add_lista_precio(data.model_dump())
        if "error" in result:
            raise HTTPException(400, result["detail"])
        return result
    except ValueError as e:
        raise HTTPException(400, str(e))


@app.put("/listas-precios/{lista_id}")
def update_lista_precio(lista_id: int, data: ListaPrecioCreate, user=Depends(get_admin_user)):
    """Actualiza una lista de precios"""
    return db.update_lista_precio(lista_id, data.model_dump())


@app.delete("/listas-precios/{lista_id}")
@limiter.limit("30/minute")
def delete_lista_precio(lista_id: int, request: Request, user=Depends(get_admin_user)):
    """Elimina una lista de precios"""
    if db.delete_lista_precio(lista_id):
        return {"message": "Lista eliminada"}
    raise HTTPException(404, "Lista no encontrada")


class PrecioEspecialCreate(BaseModel):
    producto_id: int
    precio_especial: float


@app.get("/listas-precios/{lista_id}/precios")
def get_precios_lista(lista_id: int, user=Depends(get_current_user)):
    """Obtiene todos los precios especiales de una lista"""
    return db.get_precios_lista(lista_id)


@app.post("/listas-precios/{lista_id}/precios")
def set_precio_especial(lista_id: int, data: PrecioEspecialCreate, user=Depends(get_admin_user)):
    """Establece un precio especial para un producto en una lista"""
    return db.set_precio_especial(lista_id, data.producto_id, data.precio_especial)


@app.delete("/listas-precios/{lista_id}/precios/{producto_id}")
@limiter.limit("30/minute")
def remove_precio_especial(lista_id: int, producto_id: int, request: Request, user=Depends(get_admin_user)):
    """Elimina un precio especial"""
    if db.remove_precio_especial(lista_id, producto_id):
        return {"message": "Precio especial eliminado"}
    raise HTTPException(404, "Precio especial no encontrado")


@app.post("/clientes/{cliente_id}/lista-precio")
def asignar_lista_cliente(cliente_id: int, lista_id: Optional[int] = None, user=Depends(get_current_user)):
    """Asigna una lista de precios a un cliente"""
    if db.asignar_lista_cliente(cliente_id, lista_id):
        return {"message": "Lista asignada correctamente"}
    raise HTTPException(404, "Cliente no encontrado")


@app.get("/clientes/{cliente_id}/precio/{producto_id}")
def get_precio_cliente(cliente_id: int, producto_id: int, user=Depends(get_current_user)):
    """Obtiene el precio de un producto para un cliente específico"""
    return {"precio": db.get_precio_cliente(cliente_id, producto_id)}


# === REPORTES AVANZADOS ===
@app.get("/reportes/ventas")
@limiter.limit("20/minute")
def get_reporte_ventas(
    request: Request,
    desde: Optional[str] = None, 
    hasta: Optional[str] = None, 
    user=Depends(get_current_user)
):
    """Reporte de ventas por período. Si no se especifican fechas, usa últimos 30 días"""
    from datetime import datetime, timedelta
    if not hasta:
        hasta = datetime.now().strftime("%Y-%m-%d")
    if not desde:
        desde = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    return db.get_reporte_ventas(desde, hasta)


@app.get("/reportes/inventario")
@limiter.limit("20/minute")
def get_reporte_inventario(request: Request, user=Depends(get_current_user)):
    """Reporte de estado del inventario"""
    return db.get_reporte_inventario()


@app.get("/reportes/clientes")
@limiter.limit("20/minute")
def get_reporte_clientes(request: Request, user=Depends(get_current_user)):
    """Reporte de análisis de clientes"""
    return db.get_reporte_clientes()


@app.get("/reportes/productos")
def get_reporte_productos(
    desde: Optional[str] = None,
    hasta: Optional[str] = None,
    user=Depends(get_current_user)
):
    """Reporte de productos más vendidos y tendencias"""
    from datetime import datetime, timedelta
    if not hasta:
        hasta = datetime.now().strftime("%Y-%m-%d")
    if not desde:
        desde = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    return db.get_reporte_productos(desde, hasta)


@app.get("/reportes/rendimiento")
def get_reporte_rendimiento(user=Depends(get_current_user)):
    """Reporte de rendimiento operativo"""
    return db.get_reporte_rendimiento()


@app.get("/reportes/comparativo")
def get_reporte_comparativo(user=Depends(get_current_user)):
    """Reporte comparativo por períodos"""
    return db.get_reporte_comparativo()


# === PEDIDOS RECURRENTES (Templates) ===
@app.get("/templates")
def get_templates(cliente_id: Optional[int] = None, user=Depends(get_current_user)):
    """Obtiene templates de pedidos"""
    return db.get_templates_pedido(cliente_id)


@app.get("/templates/{template_id}")
def get_template(template_id: int, user=Depends(get_current_user)):
    """Obtiene un template con sus detalles"""
    template = db.get_template_pedido(template_id)
    if not template:
        raise HTTPException(404, "Template no encontrado")
    return template


class ProductoTemplate(BaseModel):
    producto_id: int
    cantidad: float = 1
    tipo: str = "unidad"


class TemplateCreate(BaseModel):
    nombre: str
    cliente_id: Optional[int] = None
    frecuencia: Optional[str] = None
    productos: List[ProductoTemplate] = []


@app.post("/templates")
def create_template(data: TemplateCreate, user=Depends(get_current_user)):
    """Crea un nuevo template de pedido"""
    try:
        return db.add_template_pedido(data.model_dump(), user["username"])
    except ValueError as e:
        raise HTTPException(400, str(e))


@app.put("/templates/{template_id}")
def update_template(template_id: int, data: TemplateCreate, user=Depends(get_current_user)):
    """Actualiza un template de pedido"""
    try:
        return db.update_template_pedido(template_id, data.model_dump())
    except ValueError as e:
        raise HTTPException(400, str(e))


@app.delete("/templates/{template_id}")
@limiter.limit("30/minute")
def delete_template(template_id: int, request: Request, user=Depends(get_current_user)):
    """Elimina un template de pedido"""
    if db.delete_template_pedido(template_id):
        return {"message": "Template eliminado"}
    raise HTTPException(404, "Template no encontrado")


@app.post("/templates/{template_id}/ejecutar")
def ejecutar_template(template_id: int, user=Depends(get_current_user)):
    """Crea un pedido real desde un template"""
    pedido_id = db.crear_pedido_desde_template(template_id, user["username"])
    if not pedido_id:
        raise HTTPException(404, "Template no encontrado")
    return {"pedido_id": pedido_id, "message": "Pedido creado desde template"}


@app.get("/clientes/{cliente_id}/ultimo-pedido")
def get_ultimo_pedido_cliente(cliente_id: int, user=Depends(get_current_user)):
    """Obtiene el último pedido de un cliente para poder repetirlo"""
    pedido = db.get_ultimo_pedido_cliente(cliente_id)
    if not pedido:
        raise HTTPException(404, "El cliente no tiene pedidos anteriores")
    return pedido


@app.post("/clientes/{cliente_id}/repetir-pedido")
def repetir_pedido_cliente(cliente_id: int, user=Depends(get_current_user)):
    """Crea un nuevo pedido copiando el último pedido del cliente"""
    ultimo = db.get_ultimo_pedido_cliente(cliente_id)
    if not ultimo:
        raise HTTPException(404, "El cliente no tiene pedidos anteriores")
    
    # Crear nuevo pedido con los mismos productos
    from datetime import datetime
    nuevo_pedido = db.add_pedido({
        "cliente_id": cliente_id,
        "fecha": datetime.now().strftime("%Y-%m-%d"),
        "productos": [{"id": p["producto_id"], "cantidad": p["cantidad"], "tipo": p["tipo"]} 
                      for p in ultimo["productos"]]
    }, user["username"])
    return {"pedido_id": nuevo_pedido["id"], "message": "Pedido repetido correctamente"}


# === Gestión de Usuarios ===
@app.get("/usuarios")
def get_all_usuarios(user=Depends(get_admin_user)):
    """Lista todos los usuarios del sistema (solo admin)"""
    return db.get_all_usuarios()


@app.put("/usuarios/{user_id}/activar")
def activar_usuario(request: Request, user_id: int, user=Depends(get_admin_user)):
    """Activa un usuario (solo admin)"""
    ctx = get_request_context(request)
    db.set_usuario_activo(user_id, 1)
    logger.info(f"USUARIO_ACTIVATED: id={user_id} by={user['username']}")
    db.audit_log(user["username"], "ACTIVATE", "usuarios", registro_id=user_id,
                 datos_despues={"activo": True}, ip_address=ctx["ip"], user_agent=ctx["user_agent"])
    return {"message": "Usuario activado"}


@app.put("/usuarios/{user_id}/desactivar")
def desactivar_usuario(request: Request, user_id: int, user=Depends(get_admin_user)):
    """Desactiva un usuario (solo admin)"""
    ctx = get_request_context(request)
    db.set_usuario_activo(user_id, 0)
    logger.info(f"USUARIO_DEACTIVATED: id={user_id} by={user['username']}")
    db.audit_log(user["username"], "DEACTIVATE", "usuarios", registro_id=user_id,
                 datos_despues={"activo": False}, ip_address=ctx["ip"], user_agent=ctx["user_agent"])
    return {"message": "Usuario desactivado"}


@app.put("/usuarios/{user_id}/rol")
def cambiar_rol_usuario(request: Request, user_id: int, rol: str = Form(...), user=Depends(get_admin_user)):
    """Cambia el rol de un usuario (solo admin)"""
    ctx = get_request_context(request)
    if rol not in ["admin", "usuario", "vendedor"]:
        raise HTTPException(400, "Rol inválido. Use: admin, usuario o vendedor")
    db.update_usuario_rol(user_id, rol)
    logger.info(f"USUARIO_ROL_CHANGED: id={user_id} new_rol={rol} by={user['username']}")
    db.audit_log(user["username"], "CHANGE_ROL", "usuarios", registro_id=user_id,
                 datos_despues={"rol": rol}, ip_address=ctx["ip"], user_agent=ctx["user_agent"])
    return {"message": f"Rol cambiado a {rol}"}


@app.delete("/usuarios/{user_id}")
@limiter.limit("30/minute")
def eliminar_usuario(request: Request, user_id: int, user=Depends(get_admin_user)):
    """Elimina un usuario (solo admin)"""
    ctx = get_request_context(request)
    if db.delete_usuario(user_id):
        logger.info(f"USUARIO_DELETED: id={user_id} by={user['username']}")
        db.audit_log(user["username"], "DELETE", "usuarios", registro_id=user_id,
                     ip_address=ctx["ip"], user_agent=ctx["user_agent"])
        return {"message": "Usuario eliminado"}
    raise HTTPException(404, "Usuario no encontrado")


@app.put("/usuarios/{user_id}/reset-password")
@limiter.limit("10/minute")
def reset_usuario_password(
    request: Request,
    user_id: int,
    new_password: str = Form(..., min_length=8, max_length=100),
    user=Depends(get_admin_user)
):
    """
    Resetea la contraseña de un usuario (solo admin).
    Genera un nuevo hash bcrypt y actualiza la BD.
    La contraseña debe cumplir los mismos requisitos que en el registro.
    """
    ctx = get_request_context(request)
    
    # Validar fortaleza del password (mismas reglas que registro)
    is_valid, msg = validate_password_strength(new_password)
    if not is_valid:
        raise HTTPException(status_code=400, detail=msg)
    
    # Obtener usuario objetivo
    usuarios = db.get_all_usuarios()
    target_user = next((u for u in usuarios if u['id'] == user_id), None)
    
    if not target_user:
        raise HTTPException(404, "Usuario no encontrado")
    
    # Hash de la nueva contraseña
    new_hash = hash_password(new_password)
    
    # Actualizar contraseña
    if db.update_usuario_password(target_user['username'], new_hash):
        logger.info(f"PASSWORD_RESET: user_id={user_id} by={user['username']}")
        db.audit_log(
            user["username"], "RESET_PASSWORD", "usuarios",
            registro_id=user_id,
            datos_despues={"password_reset": True},
            ip_address=ctx["ip"],
            user_agent=ctx["user_agent"]
        )
        return {"message": f"Contraseña de {target_user['username']} reseteada exitosamente"}
    
    raise HTTPException(500, "Error al resetear contraseña")


# === Estadísticas y tracking ===
@app.get("/estadisticas/usuarios")
def get_estadisticas_usuarios(user=Depends(get_admin_user)):
    """Obtiene estadísticas de actividad por usuario"""
    return db.get_estadisticas_usuarios()


# === Backup Management ===
@app.post("/admin/backup")
@limiter.limit("5/hour")
def create_backup(request: Request, user=Depends(get_admin_user)):
    """
    Creates a backup of the database.
    Returns the backup file path and metadata.
    Can be triggered by cron jobs or external monitoring services.
    """
    import shutil
    from datetime import datetime
    
    ctx = get_request_context(request)
    
    # Get database path
    db_path = os.getenv("DB_PATH", "/data/ventas.db")
    if not os.path.exists(db_path):
        raise HTTPException(status_code=404, detail="Database file not found")
    
    # Create backups directory if needed
    backup_dir = os.path.join(os.path.dirname(db_path), "backups")
    os.makedirs(backup_dir, exist_ok=True)
    
    # Generate backup filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_filename = f"ventas.db.{timestamp}.bak"
    backup_path = os.path.join(backup_dir, backup_filename)
    
    try:
        # Copy database file (using SQLite backup API would be better for live DBs)
        shutil.copy2(db_path, backup_path)
        
        # Get file sizes
        original_size = os.path.getsize(db_path)
        backup_size = os.path.getsize(backup_path)
        
        logger.info(f"BACKUP_CREATED: path={backup_path} size={backup_size} by={user['username']}")
        db.audit_log(user["username"], "BACKUP", "system",
                     datos_despues={"backup_path": backup_path, "size": backup_size},
                     ip_address=ctx["ip"], user_agent=ctx["user_agent"])
        
        # Clean old backups (keep last 10)
        existing_backups = sorted([
            f for f in os.listdir(backup_dir) 
            if f.startswith("ventas.db.") and f.endswith(".bak")
        ], reverse=True)
        
        for old_backup in existing_backups[10:]:
            try:
                os.remove(os.path.join(backup_dir, old_backup))
                logger.info(f"BACKUP_CLEANED: removed old backup {old_backup}")
            except Exception:
                pass
        
        return {
            "success": True,
            "backup_path": backup_path,
            "backup_filename": backup_filename,
            "timestamp": timestamp,
            "original_size_bytes": original_size,
            "backup_size_bytes": backup_size,
            "backups_retained": min(len(existing_backups) + 1, 10)
        }
    except Exception as e:
        logger.error(f"BACKUP_FAILED: error={str(e)}")
        raise HTTPException(status_code=500, detail=f"Backup failed: {str(e)}")


@app.get("/admin/backups")
def list_backups(user=Depends(get_admin_user)):
    """List all available database backups"""
    db_path = os.getenv("DB_PATH", "/data/ventas.db")
    backup_dir = os.path.join(os.path.dirname(db_path), "backups")
    
    if not os.path.exists(backup_dir):
        return {"backups": [], "count": 0}
    
    backups = []
    for f in sorted(os.listdir(backup_dir), reverse=True):
        if f.startswith("ventas.db.") and f.endswith(".bak"):
            filepath = os.path.join(backup_dir, f)
            stat = os.stat(filepath)
            backups.append({
                "filename": f,
                "size_bytes": stat.st_size,
                "created_at": datetime.fromtimestamp(stat.st_mtime).isoformat()
            })
    
    return {"backups": backups, "count": len(backups)}


@app.get("/admin/backup/health")
def backup_health(request: Request):
    """
    Health check endpoint for external backup monitoring services.
    Can be called by UptimeRobot, cron jobs, etc. to trigger periodic backups.
    
    API key should be passed in X-API-Key header for security (not in query string).
    """
    expected_key = os.getenv("BACKUP_API_KEY")
    api_key = request.headers.get("X-API-Key")
    
    # If BACKUP_API_KEY is configured, require valid key
    if expected_key:
        if not api_key:
            raise HTTPException(status_code=401, detail="API key required in X-API-Key header")
        if api_key != expected_key:
            raise HTTPException(status_code=403, detail="Invalid API key")
    
    db_path = os.getenv("DB_PATH", "/data/ventas.db")
    backup_dir = os.path.join(os.path.dirname(db_path), "backups")
    
    # Check last backup time
    last_backup = None
    if os.path.exists(backup_dir):
        backups = [f for f in os.listdir(backup_dir) if f.endswith(".bak")]
        if backups:
            latest = max(backups, key=lambda f: os.path.getmtime(os.path.join(backup_dir, f)))
            last_backup = datetime.fromtimestamp(
                os.path.getmtime(os.path.join(backup_dir, latest))
            ).isoformat()
    
    return {
        "status": "healthy",
        "database_exists": os.path.exists(db_path),
        "database_size_bytes": os.path.getsize(db_path) if os.path.exists(db_path) else 0,
        "last_backup": last_backup,
        "backup_dir": backup_dir
    }


@app.get("/pedidos/antiguos")
def get_pedidos_antiguos(horas: int = 24, user=Depends(get_current_user)):
    """Obtiene pedidos pendientes con más de X horas de antigüedad"""
    return db.get_pedidos_antiguos(horas)


@app.get("/pedidos/{pedido_id}/historial")
def get_historial_pedido(pedido_id: int, user=Depends(get_current_user)):
    """Obtiene el historial de modificaciones de un pedido"""
    return db.get_historial_pedido(pedido_id)


def _detect_device(request: Request) -> tuple:
    """Detecta el tipo de dispositivo desde el User-Agent"""
    user_agent = request.headers.get("user-agent", "").lower()
    
    # Detectar dispositivo
    if any(x in user_agent for x in ["mobile", "android", "iphone", "ipod"]):
        if "tablet" in user_agent or "ipad" in user_agent:
            dispositivo = "tablet"
        else:
            dispositivo = "mobile"
    elif any(x in user_agent for x in ["tablet", "ipad"]):
        dispositivo = "tablet"
    else:
        dispositivo = "web"
    
    return dispositivo, request.headers.get("user-agent", "")[:500]

# === API v1 Routes (for future-proofing) ===
@api_v1.get("/health")
def health_check_v1():
    """API v1 Health check endpoint"""
    try:
        con = db.conectar()
        cur = con.cursor()
        cur.execute("SELECT 1")
        con.close()
        db_status = "ok"
    except Exception as e:
        db_status = f"error: {str(e)}"
    
    return {
        "status": "healthy" if db_status == "ok" else "degraded",
        "database": db_status,
        "api_version": "v1",
        "version": API_VERSION,
        "timestamp": datetime.now(timezone.utc).isoformat()
    }

# Mount API v1 router
app.include_router(api_v1)